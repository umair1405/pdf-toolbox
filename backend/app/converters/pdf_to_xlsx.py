# app/converters/pdf_to_excel.py

import pdfplumber
import pandas as pd
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


def group_words_into_table(words, y_tol=6, x_tol=15):
    """
    Convert PDF word positions into rows + columns automatically.
    Works better than plain extract_text().
    """

    rows = []

    for w in words:
        placed = False

        for row in rows:
            if abs(row["y"] - w["top"]) < y_tol:
                row["words"].append(w)
                placed = True
                break

        if not placed:
            rows.append({
                "y": w["top"],
                "words": [w]
            })

    rows.sort(key=lambda r: r["y"])

    output = []

    for row in rows:

        row["words"].sort(key=lambda x: x["x0"])

        cols = []
        current = ""

        prev_x = None

        for word in row["words"]:

            if prev_x is None:
                current = word["text"]

            else:
                gap = word["x0"] - prev_x

                if gap > x_tol:
                    cols.append(current)
                    current = word["text"]

                else:
                    current += " " + word["text"]

            prev_x = word["x1"]

        if current:
            cols.append(current)

        output.append(cols)

    max_cols = max(len(r) for r in output)

    for r in output:
        while len(r) < max_cols:
            r.append("")

    return output


def auto_fit(sheet):

    for col in sheet.columns:

        max_len = 0

        column = col[0].column

        for cell in col:

            try:
                value = str(cell.value)

                if len(value) > max_len:
                    max_len = len(value)

            except:
                pass

        sheet.column_dimensions[
            get_column_letter(column)
        ].width = min(max_len + 3, 40)


def pdf_to_xlsx(input_path, output_path):

    output_path = str(output_path)

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl"
    ) as writer:

        with pdfplumber.open(input_path) as pdf:

            for page_no, page in enumerate(pdf.pages):

                page_name = f"Page_{page_no+1}"

                exported = False

                # ---------
                # Method 1
                # Real table extraction
                # ---------

                tables = page.extract_tables()

                if tables:

                    for idx, table in enumerate(tables):

                        if table and len(table) > 0:

                            df = pd.DataFrame(table)

                            df.to_excel(
                                writer,
                                sheet_name=f"{page_name}_{idx+1}",
                                index=False,
                                header=False
                            )

                            exported = True

                # ---------
                # Method 2
                # Position based extraction
                # ---------

                if not exported:

                    words = page.extract_words()

                    if words:

                        rows = group_words_into_table(words)

                        df = pd.DataFrame(rows)

                        df.to_excel(
                            writer,
                            sheet_name=page_name,
                            index=False,
                            header=False
                        )

                    else:

                        text = page.extract_text()

                        if text:

                            rows = [
                                [x]
                                for x in text.split("\n")
                            ]

                            pd.DataFrame(rows).to_excel(
                                writer,
                                sheet_name=page_name,
                                index=False,
                                header=False
                            )

    workbook = load_workbook(output_path)

    for sheet in workbook.worksheets:
        auto_fit(sheet)

    workbook.save(output_path)

    return output_path